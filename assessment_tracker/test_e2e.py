"""End-to-end smoke test: generate synthetic data, process it, build report."""
import pandas as pd
import numpy as np
import os
import sys
sys.path.insert(0, ".")
from processor import process_data
from report_builder import build_report

# Generate a small synthetic CaseWorthy export
np.random.seed(42)
offices = ["Downtown", "Northside", "Eastside", "Westside", "Unassigned"]
programs = ["SNAP", "Medicaid", "TANF"]
types = ["Program Enrollment", "Program Exit", "90 Day/Annual Recert or Update"]
statuses = [
    ("Yes", "Approved Eligibility Determination"),
    ("Yes", "Pending Approval"),
    ("No", ""),
]

rows = []
for i in range(500):
    begin = pd.Timestamp("2025-07-01") + pd.Timedelta(days=np.random.randint(0, 270))
    status = statuses[np.random.choice(len(statuses), p=[0.6, 0.1, 0.3])]
    modified = begin + pd.Timedelta(days=np.random.randint(1, 60)) if status[0] == "Yes" else pd.NaT
    rows.append({
        "Client ID": np.random.randint(1000, 1200),
        "AssessmentID": np.random.randint(5000, 5400),
        "Program Name": np.random.choice(programs),
        "Type of Assessment": np.random.choice(types),
        "Begin Date": begin,
        "Assessment.LastModifiedDate": modified,
        "Assessment.BeginAssessment": begin,
        "Last Modified Date": modified,
        "Last Case Note Date Per Prog": pd.NaT,
        "Office Location": np.random.choice(offices),
        "Program Reviewed": status[0],
        "Program Review Status": status[1],
    })

# Add some exact duplicates
rows.extend(rows[:20])

test_file = "/tmp/test_caseworthy_export.xlsx"
df = pd.DataFrame(rows)
df.to_excel(test_file, index=False)
print(f"Wrote {len(df)} rows to {test_file}")

# Process
data = process_data(test_file)

# Build report
output = build_report(data, output_dir="/tmp")
print(f"\nReport generated: {output}")
assert os.path.exists(output), "Output file not created"

# Verify workbook structure
from openpyxl import load_workbook
wb = load_workbook(output)
sheets = wb.sheetnames
print(f"Sheets: {sheets}")
assert "Summary" in sheets
assert "Office Detail" in sheets
assert "Raw Data" in sheets
assert "Backlog Trend" not in sheets
assert "Monthly Activity" not in sheets

# Check summary has totals row with formulas
ws = wb["Summary"]
# Find totals row (last data row)
for row in ws.iter_rows(min_col=1, max_col=1):
    for cell in row:
        if cell.value == "TOTALS":
            totals_row = cell.row
            # Check baseline total is a SUM formula
            baseline_cell = ws.cell(row=totals_row, column=2)
            assert str(baseline_cell.value).startswith("=SUM"), f"Expected SUM formula, got {baseline_cell.value}"
            print(f"Totals row {totals_row}: baseline formula = {baseline_cell.value}")

print("\nEnd-to-end test PASSED!")
