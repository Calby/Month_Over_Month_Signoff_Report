import os
from datetime import datetime
import calendar

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

from config import (
    BASELINE_DATE, OFFICE_COL, OUTPUT_FILENAME,
    HEADER_DARK_BLUE, HEADER_MED_BLUE, HEADER_ALT_BLUE, HEADER_ALT_MED,
    HEADER_FONT_COLOR, INCREASE_FILL, DECREASE_FILL,
    OFFICE_COL_WIDTH, BASELINE_COL_WIDTH, METRIC_COL_WIDTH,
    RAW_DATA_COLUMNS,
)

# Reusable styles
_dark_fill = PatternFill(start_color=HEADER_DARK_BLUE, end_color=HEADER_DARK_BLUE, fill_type="solid")
_med_fill = PatternFill(start_color=HEADER_MED_BLUE, end_color=HEADER_MED_BLUE, fill_type="solid")
_alt_fill = PatternFill(start_color=HEADER_ALT_BLUE, end_color=HEADER_ALT_BLUE, fill_type="solid")
_alt_med_fill = PatternFill(start_color=HEADER_ALT_MED, end_color=HEADER_ALT_MED, fill_type="solid")
_white_bold = Font(color=HEADER_FONT_COLOR, bold=True)
_white_font = Font(color=HEADER_FONT_COLOR)
_bold_font = Font(bold=True)
_red_fill = PatternFill(start_color=INCREASE_FILL, end_color=INCREASE_FILL, fill_type="solid")
_green_fill = PatternFill(start_color=DECREASE_FILL, end_color=DECREASE_FILL, fill_type="solid")
_thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _month_label(year, month):
    return f"{calendar.month_abbr[month]} {year}"


def _build_summary_sheet(wb, data):
    """Sheet 1: Summary — main backlog tracker table with legend."""
    ws = wb.active
    ws.title = "Summary"

    offices = data["offices"]
    baseline = data["baseline"]
    months = data["months"]
    monthly_data = data["monthly_data"]
    num_offices = len(offices)

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=1, value="Assessment Sign-Off Backlog Tracker")
    title_cell.font = Font(size=16, bold=True)

    # Row 2: Subtitle
    gen_date = datetime.now().strftime("%B %d, %Y")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=2, column=1,
            value=f"Generated: {gen_date} | Data baseline: Aug 29, 2025 | Source: CaseWorthy Export")

    # Row 4-5: Headers
    header_row = 4
    sub_row = 5
    data_start_row = 6

    ws.cell(row=header_row, column=1, value="Office")
    ws.cell(row=sub_row, column=1, value="")
    for r in (header_row, sub_row):
        c = ws.cell(row=r, column=1)
        c.fill = _dark_fill
        c.font = _white_bold
        c.border = _thin_border

    ws.cell(row=header_row, column=2, value="Aug 29 Baseline")
    ws.cell(row=sub_row, column=2, value="Backlog")
    for r in (header_row, sub_row):
        c = ws.cell(row=r, column=2)
        c.fill = _dark_fill
        c.font = _white_bold
        c.alignment = Alignment(horizontal="center")
        c.border = _thin_border

    col = 3
    month_col_starts = {}
    sub_labels = ["New", "Signed Off", "Pending", "End Backlog", "Delta"]
    for m_idx, (year, month) in enumerate(months):
        label = _month_label(year, month)
        month_col_starts[(year, month)] = col
        # Alternate colors between months
        is_alt = m_idx % 2 == 1
        top_fill = _alt_fill if is_alt else _dark_fill
        sub_fill = _alt_med_fill if is_alt else _med_fill
        # Merge month header across 5 cols
        ws.merge_cells(start_row=header_row, start_column=col,
                       end_row=header_row, end_column=col + 4)
        h = ws.cell(row=header_row, column=col, value=label)
        h.fill = top_fill
        h.font = _white_bold
        h.alignment = Alignment(horizontal="center")
        h.border = _thin_border
        for i, sub in enumerate(sub_labels):
            sc = ws.cell(row=sub_row, column=col + i, value=sub)
            sc.fill = sub_fill
            sc.font = _white_font
            sc.alignment = Alignment(horizontal="center")
            sc.border = _thin_border
        col += 5

    # Data rows
    for idx, office in enumerate(offices):
        row = data_start_row + idx
        oc = ws.cell(row=row, column=1, value=office)
        oc.font = _bold_font
        oc.border = _thin_border
        ws.cell(row=row, column=2, value=int(baseline[office])).border = _thin_border

        prev_backlog_col = 2
        for year, month in months:
            c = month_col_starts[(year, month)]
            md = monthly_data[(year, month)]
            ws.cell(row=row, column=c, value=int(md["new"][office])).border = _thin_border
            ws.cell(row=row, column=c + 1, value=int(md["signed_off"][office])).border = _thin_border
            ws.cell(row=row, column=c + 2, value=int(md["pending"][office])).border = _thin_border
            backlog_cell = ws.cell(row=row, column=c + 3, value=int(md["eom_backlog"][office]))
            backlog_cell.border = _thin_border

            bl_col_letter = get_column_letter(c + 3)
            prev_col_letter = get_column_letter(prev_backlog_col)
            delta_cell = ws.cell(row=row, column=c + 4)
            delta_cell.value = f"={bl_col_letter}{row}-{prev_col_letter}{row}"
            delta_cell.border = _thin_border

            prev_backlog_col = c + 3

    # Conditional formatting for backlog and delta cells
    for idx, office in enumerate(offices):
        row = data_start_row + idx
        prev_backlog_val = int(baseline[office])
        for year, month in months:
            c = month_col_starts[(year, month)]
            md = monthly_data[(year, month)]
            cur_backlog_val = int(md["eom_backlog"][office])
            backlog_cell = ws.cell(row=row, column=c + 3)
            delta_cell = ws.cell(row=row, column=c + 4)
            if cur_backlog_val > prev_backlog_val:
                backlog_cell.fill = _red_fill
                delta_cell.fill = _red_fill
            elif cur_backlog_val < prev_backlog_val:
                backlog_cell.fill = _green_fill
                delta_cell.fill = _green_fill
            prev_backlog_val = cur_backlog_val

    # Totals row (Excel SUM formulas)
    totals_row = data_start_row + num_offices
    tc = ws.cell(row=totals_row, column=1, value="TOTALS")
    tc.font = Font(bold=True, size=11)
    tc.border = _thin_border

    bl_letter = get_column_letter(2)
    ws.cell(row=totals_row, column=2,
            value=f"=SUM({bl_letter}{data_start_row}:{bl_letter}{totals_row - 1})")
    ws.cell(row=totals_row, column=2).font = _bold_font
    ws.cell(row=totals_row, column=2).border = _thin_border

    for year, month in months:
        c = month_col_starts[(year, month)]
        for offset in range(5):
            col_letter = get_column_letter(c + offset)
            cell = ws.cell(row=totals_row, column=c + offset)
            cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{totals_row - 1})"
            cell.font = _bold_font
            cell.border = _thin_border

    # --- Legend / Instructions ---
    legend_start = totals_row + 3
    ws.merge_cells(start_row=legend_start, start_column=1,
                   end_row=legend_start, end_column=6)
    lh = ws.cell(row=legend_start, column=1, value="Column Definitions")
    lh.font = Font(size=13, bold=True)

    definitions = [
        ("Baseline Backlog", "Number of assessments needing sign-off as of August 29, 2025 — the starting point for tracking."),
        ("New", "Assessments with a Begin Date in that month. These are newly created assessments entering the queue."),
        ("Signed Off", "Assessments approved (signed off) during that month, based on the Assessment.LastModifiedDate."),
        ("Pending", "Assessments marked as reviewed but with a \"Pending Approval\" status — they are NOT fully signed off and still count toward the backlog."),
        ("End Backlog", "Total assessments still awaiting sign-off at the end of that month. This is the cumulative backlog."),
        ("Delta", "Change in backlog compared to the prior month. Negative values (green) = backlog is shrinking. Positive values (red) = backlog is growing."),
    ]

    for i, (term, desc) in enumerate(definitions):
        r = legend_start + 1 + i
        tc = ws.cell(row=r, column=1, value=term)
        tc.font = _bold_font
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2, value=desc)

    color_start = legend_start + len(definitions) + 2
    ws.merge_cells(start_row=color_start, start_column=1,
                   end_row=color_start, end_column=6)
    ch = ws.cell(row=color_start, column=1, value="Color Key")
    ch.font = Font(size=13, bold=True)

    gr = ws.cell(row=color_start + 1, column=1, value="   ")
    gr.fill = _green_fill
    ws.merge_cells(start_row=color_start + 1, start_column=2,
                   end_row=color_start + 1, end_column=6)
    ws.cell(row=color_start + 1, column=2,
            value="Green — Backlog decreased from the prior month (improvement)")

    rr = ws.cell(row=color_start + 2, column=1, value="   ")
    rr.fill = _red_fill
    ws.merge_cells(start_row=color_start + 2, start_column=2,
                   end_row=color_start + 2, end_column=6)
    ws.cell(row=color_start + 2, column=2,
            value="Red — Backlog increased from the prior month (falling behind)")

    # Column widths
    ws.column_dimensions["A"].width = OFFICE_COL_WIDTH
    ws.column_dimensions["B"].width = BASELINE_COL_WIDTH
    for year, month in months:
        c = month_col_starts[(year, month)]
        for offset in range(5):
            ws.column_dimensions[get_column_letter(c + offset)].width = METRIC_COL_WIDTH

    ws.freeze_panes = "B6"


def _build_trend_chart_sheet(wb, data):
    """Sheet 2: Backlog Trend line chart (top 10 offices)."""
    ws = wb.create_sheet("Backlog Trend")

    offices = data["offices"]
    baseline = data["baseline"]
    months = data["months"]
    monthly_data = data["monthly_data"]

    # Top 10 by baseline backlog
    top10 = baseline.sort_values(ascending=False).head(10).index.tolist()
    num_rows = len(top10)
    num_time_points = 1 + len(months)  # baseline + each month

    # Data table in COLUMN layout (openpyxl charts work best this way)
    # Col A = time labels, Col B = office 1, Col C = office 2, etc.
    ws.cell(row=1, column=1, value="Month")
    for i, office in enumerate(top10):
        ws.cell(row=1, column=2 + i, value=office)

    # Row 2 = baseline
    ws.cell(row=2, column=1, value="Aug 29 Baseline")
    for i, office in enumerate(top10):
        ws.cell(row=2, column=2 + i, value=int(baseline[office]))

    # Rows 3+ = each month
    for m_idx, (year, month) in enumerate(months):
        row = 3 + m_idx
        ws.cell(row=row, column=1, value=_month_label(year, month))
        for i, office in enumerate(top10):
            ws.cell(row=row, column=2 + i,
                    value=int(monthly_data[(year, month)]["eom_backlog"][office]))

    last_data_row = 2 + len(months)
    last_data_col = 1 + num_rows

    chart = LineChart()
    chart.title = "Backlog Trend by Office"
    chart.style = 10
    chart.y_axis.title = "End-of-Month Backlog"
    chart.x_axis.title = "Month"
    chart.width = 28
    chart.height = 16

    # Categories = month labels in col A
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_data_row)
    # Data = each office column (with header as series name)
    data_ref = Reference(ws, min_col=2, max_col=last_data_col,
                         min_row=1, max_row=last_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    chart_row = last_data_row + 2
    ws.add_chart(chart, f"A{chart_row}")


def _build_activity_chart_sheet(wb, data):
    """Sheet 3: Monthly Activity bar chart — New vs Signed Off."""
    ws = wb.create_sheet("Monthly Activity")

    months = data["months"]
    monthly_data = data["monthly_data"]

    # Data table
    ws.cell(row=1, column=1, value="Month")
    ws.cell(row=1, column=2, value="New Assessments")
    ws.cell(row=1, column=3, value="Signed Off")

    for i, (year, month) in enumerate(months):
        row = 2 + i
        ws.cell(row=row, column=1, value=_month_label(year, month))
        ws.cell(row=row, column=2, value=int(monthly_data[(year, month)]["new"].sum()))
        ws.cell(row=row, column=3, value=int(monthly_data[(year, month)]["signed_off"].sum()))

    last_data_row = 1 + len(months)

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "New vs Signed Off — All Offices Combined"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Month"
    chart.width = 24
    chart.height = 14

    cats = Reference(ws, min_col=1, min_row=2, max_row=last_data_row)
    data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=last_data_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    chart_row = last_data_row + 2
    ws.add_chart(chart, f"A{chart_row}")


def _build_detail_sheet(wb, data):
    """Sheet 4: Office Detail Table."""
    ws = wb.create_sheet("Office Detail")

    detail = data["detail"]
    headers = ["Office", "Total Assessments", "Current Backlog", "Signed Off",
               "Pending Review", "% Signed Off"]

    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = _dark_fill
        c.font = _white_bold
        c.alignment = Alignment(horizontal="center")
        c.border = _thin_border

    for r_idx, (office, row) in enumerate(detail.iterrows(), start=2):
        ws.cell(row=r_idx, column=1, value=office).border = _thin_border
        ws.cell(row=r_idx, column=2, value=int(row["total_assessments"])).border = _thin_border
        ws.cell(row=r_idx, column=3, value=int(row["current_backlog"])).border = _thin_border
        ws.cell(row=r_idx, column=4, value=int(row["signed_off_count"])).border = _thin_border
        ws.cell(row=r_idx, column=5, value=int(row["pending_review_count"])).border = _thin_border
        pct_cell = ws.cell(row=r_idx, column=6, value=row["pct_signed_off"] / 100)
        pct_cell.number_format = "0.0%"
        pct_cell.border = _thin_border

    ws.column_dimensions["A"].width = OFFICE_COL_WIDTH
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 18
    ws.freeze_panes = "A2"


def _build_raw_data_sheet(wb, data):
    """Sheet 5: Raw Data — only report-relevant columns."""
    ws = wb.create_sheet("Raw Data")

    raw = data["raw"]
    # Only include columns relevant to this report
    columns = [c for c in RAW_DATA_COLUMNS if c in raw.columns]
    filtered = raw[columns]

    for i, h in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = _dark_fill
        c.font = _white_bold
        c.border = _thin_border

    for r_idx, (_, row) in enumerate(filtered.iterrows(), start=2):
        for c_idx, col in enumerate(columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = ""
            elif isinstance(val, pd.Timestamp):
                val = val.strftime("%Y-%m-%d %H:%M:%S") if val.hour else val.strftime("%Y-%m-%d")
            ws.cell(row=r_idx, column=c_idx, value=val)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_report(data: dict, output_dir: str | None = None) -> str:
    """Build the full Excel workbook and return the output path."""
    wb = Workbook()

    _build_summary_sheet(wb, data)
    _build_trend_chart_sheet(wb, data)
    _build_activity_chart_sheet(wb, data)
    _build_detail_sheet(wb, data)
    _build_raw_data_sheet(wb, data)

    datestamp = datetime.now().strftime("%Y%m%d")
    filename = f"Assessment_SignOff_Tracker_{datestamp}.xlsx"
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        path = os.path.join(output_dir, filename)
    else:
        path = filename

    wb.save(path)
    return path
